#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Consolidate Wuerzburg Stadtbezirk social/demographic open-data into one .xlsx.
Source: Stadt Wuerzburg - opendata.wuerzburg.de (Explore API v2.1), licence dl-de/by-2-0.
All inputs are the raw /exports/csv dumps already saved under data/.
"""
import pandas as pd, numpy as np, os, datetime, time, urllib.request

DATA="data"
API="https://opendata.wuerzburg.de/api/explore/v2.1/catalog/datasets/"
FETCH_IDS=[
 "stadtbezirke","sozialmonitoring-betreuungsplaetze-fur-kinder",
 "stadtbezirke_hauptwohnsitz","stadtbezirke_hauptwohnsitz_weiblich","stadtbezirke_hauptwohnsitz_altersgruppen",
 "stadtbezirke_wohnberechtigte","stadtbezirke_wohnberechtigte_weiblich",
 "stadtbezirke_auslaender","stadtbezirke_auslaender_weiblich","stadtbezirke_auslaender_durchschnittsalter",
 "stadtbezirke_deutsche_mit_migrationshintergrund","stadtbezirke_deutsche_mit_migrationshintergrund_weiblich",
 "stadtbezirke_deutsche_mit_migrationshintergrund_durchschnittsalter",
 "stadtbezirke_durchschnittsalter","stadtbezirke_medianalter",
 "stadtbezirke_jugendquotient","stadtbezirke_altenquotient","stadtbezirke_abhaengigkeitsquotient","stadtbezirke_greyingindex",
 "stadtbezirke_haushalte","stadtbezirke_haushalte_durchschnittsgroesse","stadtbezirke_haushalte_mit_kindern_unter_18jahre",
 "stadtbezirke_bewegung_geburten","stadtbezirke_bewegung_sterbefaelle","stadtbezirke_bewegung_zuzuege","stadtbezirke_bewegung_wegzuege",
]

def ensure_data():
    """Download each dataset's CSV export once into DATA/ (skips files already present)."""
    os.makedirs(DATA,exist_ok=True)
    for did in FETCH_IDS:
        fp=f"{DATA}/{did}.csv"
        if os.path.exists(fp) and os.path.getsize(fp)>0: continue
        url=API+did+"/exports/csv"
        for a in range(6):
            try:
                req=urllib.request.Request(url,headers={"User-Agent":"wue-stadtbezirke/1.0"})
                open(fp,"wb").write(urllib.request.urlopen(req,timeout=120).read())
                print("fetched",did); break
            except Exception as e:
                print("retry",did,e); time.sleep(2*(a+1))

if os.environ.get("FETCH","1")=="1":
    ensure_data()
ATTR="Quelle: Stadt Würzburg – opendata.wuerzburg.de (Lizenz dl-de/by-2-0)"
PULL_DATE=datetime.date.today().isoformat()

# ---------------------------------------------------------------- key handling
def fold(s):
    if s is None or (isinstance(s,float) and np.isnan(s)): return None
    s=str(s).strip()
    return (s.replace("ä","ae").replace("ö","oe").replace("ü","ue")
             .replace("Ä","Ae").replace("Ö","Oe").replace("Ü","Ue")
             .replace("ß","ss"))

# canonical display name (with umlauts, from the 'stadtbezirke' spine) + nummer
SPINE=[("01","Altstadt"),("02","Zellerau"),("03","Dürrbachtal"),("04","Grombühl"),
       ("05","Lindleinsmühle"),("06","Frauenland"),("07","Sanderau"),("08","Heidingsfeld"),
       ("09","Heuchelhof"),("10","Steinbachtal"),("11","Versbach"),("12","Lengfeld"),
       ("13","Rottenbauer")]
FOLD2CANON={fold(n):(nr,n) for nr,n in SPINE}
GESAMT_FOLDS={"Wuerzburg","Wurzburg","Gesamt","Gesamtstadt","Stadt Wuerzburg"}

def resolve(name):
    """return (nummer, canonical) or ('GES','Gesamtstadt') or (None,None)"""
    f=fold(name)
    if f is None: return (None,None)
    if f in FOLD2CANON: return FOLD2CANON[f]
    if f in GESAMT_FOLDS: return ("GES","Gesamtstadt")
    return (None,None)

def yr(series):
    return pd.to_numeric(series.astype(str).str[:4],errors="coerce").astype("Int64")

def load(did):
    return pd.read_csv(f"{DATA}/{did}.csv",sep=";")

# ---------------------------------------------------------------- metric registry
# scalar datasets: (dataset_id, value_column, metric_key, label, kind, definition)
SCALAR=[
 ("stadtbezirke_hauptwohnsitz","wert","einwohner_hw","Einwohner (Hauptwohnsitz)","count",
  "Melderechtlich registrierte Einwohner am Ort der Hauptwohnung (Mehrfachwohnsitze einfach gezählt)."),
 ("stadtbezirke_hauptwohnsitz_weiblich","wert","einwohner_hw_w","Einwohner (HW) weiblich","count","HW-Einwohner weiblich."),
 ("stadtbezirke_wohnberechtigte","wert","wohnberechtigte","Wohnberechtigte (Haupt+Neben)","count","Wohnberechtigte Bevölkerung (Haupt- und Nebenwohnsitz)."),
 ("stadtbezirke_wohnberechtigte_weiblich","wert","wohnberechtigte_w","Wohnberechtigte weiblich","count","Wohnberechtigte weiblich."),
 ("stadtbezirke_auslaender","wert","auslaender","Ausländer (Anzahl)","count",
  "Personen mit ausschließlich ausländischer oder ungeklärter Staatsangehörigkeit, inkl. Staatenlose; HW."),
 ("stadtbezirke_auslaender_weiblich","wert","auslaender_w","Ausländer weiblich","count","Ausländerinnen, HW."),
 ("stadtbezirke_auslaender_durchschnittsalter","wert","auslaender_durchschnittsalter","Ausländer Durchschnittsalter","avg","Mittleres Alter der Ausländer."),
 ("stadtbezirke_deutsche_mit_migrationshintergrund","wert","migrationshintergrund","Deutsche mit Migrationshintergrund (Anzahl)","count",
  "Eingebürgerte, im Ausland geborene Deutsche, Aussiedler; zzgl. unter-18-Jährige mit ≥1 Elternteil mit Migrationshintergrund; HW."),
 ("stadtbezirke_deutsche_mit_migrationshintergrund_weiblich","wert","migrationshintergrund_w","Deutsche m. Migrationshintergrund weiblich","count","Deutsche mit Migrationshintergrund, weiblich."),
 ("stadtbezirke_deutsche_mit_migrationshintergrund_durchschnittsalter","wert","migrationshintergrund_durchschnittsalter","Migrationshintergrund Durchschnittsalter","avg","Mittleres Alter der Deutschen mit Migrationshintergrund."),
 ("stadtbezirke_durchschnittsalter","wert","durchschnittsalter","Durchschnittsalter (gesamt)","avg","Mittleres Alter aller HW-Einwohner."),
 ("stadtbezirke_medianalter","wert","medianalter","Medianalter","median","Medianalter aller HW-Einwohner."),
 ("stadtbezirke_jugendquotient","wert","jugendquotient","Jugendquotient","ratio","Unter-20-Jährige je 100 Personen im Erwerbsalter (20–64)."),
 ("stadtbezirke_altenquotient","wert","altenquotient","Altenquotient","ratio","65+ je 100 Personen im Erwerbsalter (20–64)."),
 ("stadtbezirke_abhaengigkeitsquotient","wert","abhaengigkeitsquotient","Abhängigkeitsquotient","ratio","(Unter-20 + 65+) je 100 Personen im Erwerbsalter."),
 ("stadtbezirke_greyingindex","wert","greyingindex","Greying-Index","ratio","Verhältnis Alte zu Jungen (Alterung)."),
 ("stadtbezirke_haushalte","value","haushalte","Haushalte (Anzahl)","count","Anzahl der Haushalte."),
 ("stadtbezirke_haushalte_durchschnittsgroesse","haushaltsgrosse","haushalte_groesse","Haushalte Durchschnittsgröße","avg","Durchschnittliche Personen je Haushalt."),
 ("stadtbezirke_haushalte_mit_kindern_unter_18jahre","anzahl_haushalte","haushalte_mit_kindern_u18","Haushalte mit Kindern <18","count","Haushalte mit mindestens einem Kind unter 18 Jahren."),
 ("stadtbezirke_bewegung_geburten","wert","geburten","Geburten","count","Lebendgeburten im Jahr."),
 ("stadtbezirke_bewegung_sterbefaelle","wert","sterbefaelle","Sterbefälle","count","Sterbefälle im Jahr."),
 ("stadtbezirke_bewegung_zuzuege","wert","zuzuege","Zuzüge","count","Zuzüge im Jahr."),
 ("stadtbezirke_bewegung_wegzuege","wert","wegzuege","Wegzüge","count","Wegzüge im Jahr."),
]

records=[]   # long: nummer, stadtbezirk, jahr, metric, value, label, kind, source, art
catalog=[]   # metric documentation rows

def add_long(nr,sb,jahr,metric,value):
    records.append(dict(nummer=nr,stadtbezirk=sb,jahr=jahr,metric=metric,value=value))

def invalid_count_years(df,col):
    """Return implausible years for COUNT data. Two placeholder patterns seen in 2025:
    (a) yearly all-Bezirk sum collapses to <50% of the median annual sum;
    (b) a year's per-Bezirk vector exactly duplicates the previous year's."""
    ys=df.groupby("jahr")[col].sum()
    if len(ys)<3: return set()
    bad=set(ys[ys<0.5*ys.median()].index.tolist())
    years=sorted(df["jahr"].dropna().unique())
    for prev,cur in zip(years,years[1:]):
        a=df[df["jahr"]==prev].set_index("nummer")[col]
        b=df[df["jahr"]==cur].set_index("nummer")[col]
        common=a.index.intersection(b.index)
        if len(common)>=10 and (a[common]==b[common]).all():
            bad.add(cur)
    return bad

def pull_scalar(did,col,metric,label,kind,defi):
    df=load(did)
    if col not in df.columns:
        # fallback: single non-key column
        cand=[c for c in df.columns if c not in ("stadtbezirk","jahr")]
        col=cand[0]
    df=df[["stadtbezirk","jahr",col]].copy()
    df["jahr"]=yr(df["jahr"])
    res=df["stadtbezirk"].map(resolve)
    df["nummer"]=res.map(lambda x:x[0]); df["canon"]=res.map(lambda x:x[1])
    orphan=df[df["nummer"].isna()]
    df=df.dropna(subset=["nummer","jahr"])
    dropped=set()
    if kind=="count":
        dropped=invalid_count_years(df,col)
        if dropped: df=df[~df["jahr"].isin(dropped)]
    yrs=sorted(df["jahr"].dropna().unique().tolist())
    for _,r in df.iterrows():
        if pd.notna(r[col]):
            add_long(r["nummer"],r["canon"],int(r["jahr"]),metric,float(r[col]))
    hinw=[]
    if not orphan.empty: hinw.append(f"{len(orphan)} Zeilen ohne gültigen Stadtbezirk verworfen")
    if dropped: hinw.append(f"Jahr(e) {sorted(int(d) for d in dropped)} als unvollständig/Platzhalter entfernt (Summenkollaps <50% Median oder exakte Vorjahres-Dublette)")
    catalog.append(dict(metric=metric,label=label,source_dataset=did,art="nativ",kind=kind,
        jahre=f"{yrs[0]}–{yrs[-1]}" if yrs else "—",latest=yrs[-1] if yrs else None,
        definition=defi,hinweis="; ".join(hinw)))

for did,col,m,lab,kind,defi in SCALAR:
    pull_scalar(did,col,m,lab,kind,defi)

# ---------------------------------------------------------------- Kinderanteil from age groups
NAMED=['0-2','3-5','6-9','10-14','15-17','18-24','25-29','30-44','45-59','60-64','65-74','75-84','85 und älter']
ag=load("stadtbezirke_hauptwohnsitz_altersgruppen")
ag=ag.dropna(subset=["stadtbezirk","jahr"])          # drops the null-key 'Column*' artifact rows
ag=ag[ag["altersgruppe"].isin(NAMED)].copy()
ag["jahr"]=yr(ag["jahr"])
res=ag["stadtbezirk"].map(resolve)
ag["nummer"]=res.map(lambda x:x[0]); ag["canon"]=res.map(lambda x:x[1])
ag=ag.dropna(subset=["nummer"])
BANDS={"kinder_u3":["0-2"],"kinder_u6":["0-2","3-5"],"kinder_u18":["0-2","3-5","6-9","10-14","15-17"]}
ag_yrs=sorted(ag["jahr"].dropna().unique().tolist())
for key,bands in BANDS.items():
    sub=ag[ag["altersgruppe"].isin(bands)]
    g=sub.groupby(["nummer","canon","jahr"])["wert"].sum().reset_index()
    for _,r in g.iterrows():
        add_long(r["nummer"],r["canon"],int(r["jahr"]),key,float(r["wert"]))
catalog += [
 dict(metric="kinder_u3",label="Kinder unter 3 Jahren (Anzahl)",source_dataset="stadtbezirke_hauptwohnsitz_altersgruppen",
      art="abgeleitet (Σ Altersband 0–2)",kind="count",jahre=f"{ag_yrs[0]}–{ag_yrs[-1]}",latest=ag_yrs[-1],
      definition="Summe Altersgruppe 0–2 Jahre (HW).",hinweis="Altersgruppen-Datensatz endet 2023; 2024/25 enthielten nur korrupte 'Column*'-Artefaktzeilen (null-Keys), verworfen."),
 dict(metric="kinder_u6",label="Kinder unter 6 Jahren (Anzahl)",source_dataset="stadtbezirke_hauptwohnsitz_altersgruppen",
      art="abgeleitet (Σ 0–2,3–5)",kind="count",jahre=f"{ag_yrs[0]}–{ag_yrs[-1]}",latest=ag_yrs[-1],
      definition="Summe Altersgruppen 0–2 und 3–5 Jahre (HW).",hinweis="Datensatz endet 2023."),
 dict(metric="kinder_u18",label="Kinder/Jugendliche unter 18 Jahren (Anzahl)",source_dataset="stadtbezirke_hauptwohnsitz_altersgruppen",
      art="abgeleitet (Σ 0–17)",kind="count",jahre=f"{ag_yrs[0]}–{ag_yrs[-1]}",latest=ag_yrs[-1],
      definition="Summe Altersgruppen 0–2,3–5,6–9,10–14,15–17 (HW).",hinweis="Datensatz endet 2023."),
]

# ---------------------------------------------------------------- Childcare (sozialmonitoring)
cc=load("sozialmonitoring-betreuungsplaetze-fur-kinder")
cc["jahr"]=yr(cc["jahr"])
THEMES={"Betreuungsplätze für Kinder unter 3 Jahren - absolut":("betreuungsplaetze_u3","Betreuungsplätze Kinder <3 (absolut)"),
        "Betreuungsplätze für Kinder von 3 bis 6 Jahren - absolut":("betreuungsplaetze_3bis6","Betreuungsplätze Kinder 3–6 (absolut)")}
cc_total={}  # (metric,jahr)->value from the source 'Würzburg' total row
for theme,(mkey,lab) in THEMES.items():
    sub=cc[cc["thema"]==theme]
    yrs=sorted(sub["jahr"].dropna().unique().tolist())
    for _,r in sub.iterrows():
        nr,canon=resolve(r["stadtbezirk"])
        if nr is None or pd.isna(r["wert"]): continue
        if nr=="GES":
            cc_total[(mkey,int(r["jahr"]))]=float(r["wert"]); continue
        add_long(nr,canon,int(r["jahr"]),mkey,float(r["wert"]))
    catalog.append(dict(metric=mkey,label=lab,source_dataset="sozialmonitoring-betreuungsplaetze-fur-kinder",
        art="nativ (anderer Datensatz!)",kind="count",jahre=f"{yrs[0]}–{yrs[-1]}",latest=yrs[-1],
        definition="Belegte/verfügbare Kinderbetreuungsplätze (absolut) je Stadtbezirk, Sozialmonitoring.",
        hinweis="NICHT aus stadtbezirke_*: Quelle ist Datensatz 'sozialmonitoring-betreuungsplaetze-fur-kinder', aber auf Stadtbezirks-Ebene und enthält eine eigene Gesamtstadt-Zeile ('Würzburg')."))

long=pd.DataFrame(records)

# ---------------------------------------------------------------- Gesamtstadt aggregation
# counts -> sum; averages/median/ratio -> not summable (leave out, weighted age handled below)
KIND={c["metric"]:c["kind"] for c in catalog}
count_metrics=[m for m,k in KIND.items() if k=="count"]
ges_rows=[]
for m in count_metrics:
    sub=long[long["metric"]==m]
    g=sub.groupby("jahr")["value"].sum().reset_index()
    for _,r in g.iterrows():
        ges_rows.append(dict(nummer="GES",stadtbezirk="Gesamtstadt",jahr=int(r["jahr"]),metric=m,value=float(r["value"])))
# childcare: prefer the source's own Würzburg total where present
for (m,j),v in cc_total.items():
    ges_rows=[r for r in ges_rows if not (r["metric"]==m and r["jahr"]==j and r["nummer"]=="GES")]
    ges_rows.append(dict(nummer="GES",stadtbezirk="Gesamtstadt",jahr=j,metric=m,value=v))
# population-weighted average age for Gesamtstadt (durchschnittsalter).
# Where a year's HW population was dropped as invalid (e.g. 2025), fall back to the
# latest valid population weights (age structure is stable year-on-year).
pop=long[long["metric"]=="einwohner_hw"][["nummer","jahr","value"]].rename(columns={"value":"pop"})
latest_pop_year=int(pop["jahr"].max())
latest_w=pop[pop["jahr"]==latest_pop_year][["nummer","pop"]]
for am in ["durchschnittsalter"]:
    base=long[long["metric"]==am]
    for j in sorted(base["jahr"].unique()):
        dj=base[base["jahr"]==j]
        w=pop[pop["jahr"]==j][["nummer","pop"]]
        approx=False
        if w["pop"].sum()==0 or len(w)==0:
            w=latest_w; approx=True
        m=dj.merge(w,on="nummer",how="inner")
        if m["pop"].sum()==0: continue
        ges_rows.append(dict(nummer="GES",stadtbezirk="Gesamtstadt",jahr=int(j),metric=am,
            value=round(float(np.average(m["value"],weights=m["pop"])),2)))
long=pd.concat([long,pd.DataFrame(ges_rows)],ignore_index=True)

# ---------------------------------------------------------------- Derived rates (per matching sb+jahr)
popw=long[long["metric"]=="einwohner_hw"][["nummer","jahr","value"]].rename(columns={"value":"pop"})
DERIVED=[("auslaender","auslaender_anteil_pct","Ausländeranteil (%)",
          "Ausländer ÷ Einwohner(HW) × 100, gleicher Stadtbezirk & Jahr."),
         ("migrationshintergrund","migrationshintergrund_anteil_pct","Anteil Deutsche m. Migrationshintergrund (%)",
          "Deutsche m. Migrationshintergrund ÷ Einwohner(HW) × 100, gleicher Stadtbezirk & Jahr."),
         ("kinder_u3","kinder_u3_anteil_pct","Kinderanteil unter 3 (%)","Kinder<3 ÷ Einwohner(HW) × 100 (gleiches Jahr, 2011–2023)."),
         ("kinder_u6","kinder_u6_anteil_pct","Kinderanteil unter 6 (%)","Kinder<6 ÷ Einwohner(HW) × 100 (gleiches Jahr, 2011–2023)."),
         ("kinder_u18","kinder_u18_anteil_pct","Kinderanteil unter 18 (%)","Kinder<18 ÷ Einwohner(HW) × 100 (gleiches Jahr, 2011–2023).")]
der_rows=[]
for base,mkey,lab,defi in DERIVED:
    sub=long[long["metric"]==base].merge(popw,on=["nummer","jahr"],how="inner")
    sub=sub[sub["pop"]>0]
    for _,r in sub.iterrows():
        der_rows.append(dict(nummer=r["nummer"],stadtbezirk=r["stadtbezirk"],jahr=int(r["jahr"]),
                             metric=mkey,value=round(r["value"]/r["pop"]*100,2)))
    yrs=sorted(sub["jahr"].unique().tolist())
    catalog.append(dict(metric=mkey,label=lab,source_dataset=f"abgeleitet: {base} ÷ stadtbezirke_hauptwohnsitz",
        art="abgeleitet (Rate)",kind="rate",jahre=f"{yrs[0]}–{yrs[-1]}" if yrs else "—",latest=yrs[-1] if yrs else None,
        definition=defi,hinweis="Nenner = Einwohner Hauptwohnsitz desselben Stadtbezirks & Jahres."))
long=pd.concat([long,pd.DataFrame(der_rows)],ignore_index=True)

# documented GAP rows (metrics requested but not available)
catalog.append(dict(metric="—",label="(Hinweis) Kinderbetreuungsplätze als RATE / Versorgungsquote",
    source_dataset="—",art="nicht verfügbar",kind="—",jahre="—",latest=None,
    definition="Eine Betreuungsquote (Plätze ÷ Kinder der Altersgruppe) ist nicht vorberechnet vorhanden.",
    hinweis="Absolute Platzzahlen sind vorhanden (siehe betreuungsplaetze_*). Eine Quote könnte aus Plätzen ÷ kinder_u3 bzw. (kinder_u6−kinder_u3) berechnet werden, wurde hier aber NICHT abgeleitet, da Altersbänder (Ende 2023) und Platzzahlen (bis 2025) jahresverschieden sind."))

long.to_parquet("long.parquet") if False else long.to_csv("long_all.csv",index=False)
cat=pd.DataFrame(catalog)
print("long rows:",len(long)," metrics:",long['metric'].nunique()," catalog rows:",len(cat))
print("years overall:",int(long['jahr'].min()),"-",int(long['jahr'].max()))

# ================================================================ EXCEL
from openpyxl import Workbook
from openpyxl.styles import Font,Alignment,PatternFill,Border,Side
from openpyxl.utils import get_column_letter

ORDER=["nummer","stadtbezirk","einwohner_hw","einwohner_hw_w","wohnberechtigte",
 "auslaender","auslaender_w","auslaender_anteil_pct","auslaender_durchschnittsalter",
 "migrationshintergrund","migrationshintergrund_w","migrationshintergrund_anteil_pct","migrationshintergrund_durchschnittsalter",
 "durchschnittsalter","medianalter",
 "kinder_u3","kinder_u3_anteil_pct","kinder_u6","kinder_u6_anteil_pct","kinder_u18","kinder_u18_anteil_pct",
 "betreuungsplaetze_u3","betreuungsplaetze_3bis6",
 "haushalte","haushalte_groesse","haushalte_mit_kindern_u18",
 "jugendquotient","altenquotient","abhaengigkeitsquotient","greyingindex",
 "geburten","sterbefaelle","zuzuege","wegzuege"]
LAB={c["metric"]:c["label"] for c in catalog}
LATEST={c["metric"]:c["latest"] for c in catalog}

# Sheet1 Konsolidiert: latest year per metric, header carries the year
def konsolidiert():
    sbs=[("01","Altstadt"),("02","Zellerau"),("03","Dürrbachtal"),("04","Grombühl"),
         ("05","Lindleinsmühle"),("06","Frauenland"),("07","Sanderau"),("08","Heidingsfeld"),
         ("09","Heuchelhof"),("10","Steinbachtal"),("11","Versbach"),("12","Lengfeld"),
         ("13","Rottenbauer"),("GES","Gesamtstadt")]
    rows=[]
    for nr,canon in sbs:
        row={"nummer":nr,"stadtbezirk":canon}
        for m in ORDER[2:]:
            ly=LATEST.get(m)
            if ly is None: row[m]=None; continue
            v=long[(long["metric"]==m)&(long["nummer"]==nr)&(long["jahr"]==ly)]["value"]
            row[m]=(round(float(v.iloc[0]),2) if len(v) else None)
        rows.append(row)
    return pd.DataFrame(rows)[ORDER]

K=konsolidiert()

# styles
HFILL=PatternFill("solid",fgColor="1F4E78"); HFONT=Font(bold=True,color="FFFFFF",size=10)
GFILL=PatternFill("solid",fgColor="D9E1F2")
thin=Side(style="thin",color="BFBFBF"); BORD=Border(thin,thin,thin,thin)
wb=Workbook(); wb.remove(wb.active)

def style_header(ws,ncol,rowi=1):
    for c in range(1,ncol+1):
        cell=ws.cell(row=rowi,column=c); cell.fill=HFILL; cell.font=HFONT
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); cell.border=BORD

def autowidth(ws,maxw=42):
    for ci in range(1,ws.max_column+1):
        L=get_column_letter(ci)
        m=max((len(str(ws.cell(r,ci).value)) for r in range(1,ws.max_row+1)
               if ws.cell(r,ci).value is not None),default=8)
        ws.column_dimensions[L].width=min(max(m+2,9),maxw)

# --- Sheet 1
ws=wb.create_sheet("Konsolidiert")
ws.append([f"Würzburg – Stadtbezirke: Sozial-/Demografiedaten konsolidiert (jeweils aktuellstes Jahr je Kennzahl)"])
ws.append([ATTR+f" | Datenabruf: {PULL_DATE}"])
ws.append([])
hdr=["Nr.","Stadtbezirk"]+[f"{LAB.get(m,m)}  [{LATEST.get(m)}]" for m in ORDER[2:]]
ws.append(hdr)
hrow=ws.max_row
for _,r in K.iterrows():
    ws.append([r[c] for c in ORDER])
ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=len(hdr))
ws.cell(1,1).font=Font(bold=True,size=12)
ws.cell(2,1).font=Font(italic=True,size=9,color="555555")
style_header(ws,len(hdr),hrow)
for rr in range(hrow+1,ws.max_row+1):
    if ws.cell(rr,2).value=="Gesamtstadt":
        for c in range(1,len(hdr)+1): ws.cell(rr,c).fill=GFILL; ws.cell(rr,c).font=Font(bold=True)
ws.freeze_panes="C"+str(hrow+1)
autowidth(ws)

# --- Sheet 2 Zeitreihe (long tidy)
ws2=wb.create_sheet("Zeitreihe")
zt=long.copy()
zt["label"]=zt["metric"].map(lambda m:LAB.get(m,m))
zt=zt[["nummer","stadtbezirk","jahr","metric","label","value"]].sort_values(["metric","jahr","nummer"])
ws2.append(["Nr.","Stadtbezirk","Jahr","Kennzahl-ID","Kennzahl","Wert"])
style_header(ws2,6)
for _,r in zt.iterrows():
    ws2.append([r["nummer"],r["stadtbezirk"],int(r["jahr"]),r["metric"],r["label"],
                round(float(r["value"]),2) if pd.notna(r["value"]) else None])
ws2.freeze_panes="A2"; autowidth(ws2,38)

# --- Sheet 3 Datenkatalog
ws3=wb.create_sheet("Datenkatalog")
ws3.append(["Konsolidierte Kennzahlen – Herkunft, Definition, Verfügbarkeit"])
ws3.cell(1,1).font=Font(bold=True,size=12); ws3.append([])
cols=["Kennzahl-ID","Bezeichnung","Quell-Datensatz (dataset_id)","Art","Typ","Jahre","Aktuellstes Jahr","Definition","Hinweis / Lücke"]
ws3.append(cols); hrow3=ws3.max_row
style_header(ws3,len(cols),hrow3)
catorder={m:i for i,m in enumerate(ORDER)}
catdf=pd.DataFrame(catalog)
catdf["o"]=catdf["metric"].map(lambda m:catorder.get(m,999))
catdf=catdf.sort_values(["o","metric"])
for _,r in catdf.iterrows():
    ws3.append([r["metric"],r["label"],r["source_dataset"],r["art"],r["kind"],r["jahre"],
                (int(r["latest"]) if pd.notna(r["latest"]) else "—"),r["definition"],r["hinweis"]])
for rr in range(hrow3+1,ws3.max_row+1):
    for c in range(1,len(cols)+1):
        ws3.cell(rr,c).alignment=Alignment(vertical="top",wrap_text=True); ws3.cell(rr,c).border=BORD
ws3.column_dimensions["A"].width=34; ws3.column_dimensions["B"].width=34
ws3.column_dimensions["C"].width=46; ws3.column_dimensions["H"].width=60; ws3.column_dimensions["I"].width=60
for L in ["D","E","F","G"]: ws3.column_dimensions[L].width=16
ws3.freeze_panes="A"+str(hrow3+1)

# --- Sheet 4 Coverage (metric x stadtbezirk, latest year, present/missing)
ws4=wb.create_sheet("Coverage")
ws4.append(["Abdeckung je Kennzahl × Stadtbezirk (aktuellstes Jahr der Kennzahl). ✓=vorhanden · –=fehlt"])
ws4.cell(1,1).font=Font(bold=True,size=11); ws4.append([])
sb13=[s for _,s in SPINE]
ws4.append(["Kennzahl","Jahr"]+sb13); hrow4=ws4.max_row
style_header(ws4,2+len(sb13),hrow4)
GREEN=PatternFill("solid",fgColor="C6EFCE"); RED=PatternFill("solid",fgColor="FFC7CE")
for m in ORDER[2:]:
    ly=LATEST.get(m)
    line=[LAB.get(m,m),(int(ly) if ly else "—")]
    present={}
    if ly is not None:
        sub=long[(long["metric"]==m)&(long["jahr"]==ly)]
        present=set(sub["stadtbezirk"])
    for s in sb13:
        line.append("✓" if s in present else "–")
    ws4.append(line)
    rr=ws4.max_row
    for i,s in enumerate(sb13):
        cell=ws4.cell(rr,3+i); cell.alignment=Alignment(horizontal="center")
        cell.fill=GREEN if cell.value=="✓" else RED
ws4.column_dimensions["A"].width=42
ws4.freeze_panes="C"+str(hrow4+1)

# --- Sheet 5 Quellen & Methodik
ws5=wb.create_sheet("Quellen_Methodik")
notes=[
 ("Titel","Würzburg Stadtbezirke – konsolidierte Sozial- und Demografiedaten"),
 ("Quelle / Lizenz",ATTR),
 ("Portal / API","https://opendata.wuerzburg.de  ·  Explore API v2.1  ·  /catalog/datasets/{id}/exports/csv"),
 ("Datenabruf am",PULL_DATE),
 ("Geografische Ebene","Stadt Würzburg, 13 Stadtbezirke (Dürrbachtal als eine Einheit). Keine Stadtteil-/Landkreisdaten."),
 ("Join-Schlüssel","Stadtbezirk. Namensvarianten normalisiert (Umlaut ↔ ae/oe/ue): z. B. 'Grombuehl'→'Grombühl', 'Duerrbachtal'→'Dürrbachtal'."),
 ("Bevölkerungsbegriff","Melderechtlich registrierte Einwohner am Ort der Hauptwohnung (HW); Mehrfachwohnsitze einfach gezählt."),
 ("Kinderanteil – Bänder","u3 = Altersgruppe 0–2; u6 = 0–2 + 3–5; u18 = 0–2,3–5,6–9,10–14,15–17. Anteil = Kinder ÷ HW-Einwohner × 100 (gleiches Jahr)."),
 ("Abgeleitete Raten","Ausländeranteil, Migrationshintergrund-Anteil, Kinderanteile = Zähler ÷ HW-Einwohner desselben Stadtbezirks & Jahres × 100."),
 ("Gesamtstadt-Zeile","Zählwerte = Summe der 13 Stadtbezirke; Durchschnittsalter = bevölkerungsgewichtetes Mittel; Quotienten/Median nicht aggregierbar (leer). Betreuungsplätze: Gesamtwert aus der portaleigenen 'Würzburg'-Zeile."),
 ("Datenqualität: Altersgruppen","stadtbezirke_hauptwohnsitz_altersgruppen enthält 2126 korrupte Zeilen ohne Stadtbezirk/Jahr (Label 'Column 1'…'Column 15'); verworfen. Saubere Altersdaten nur 2011–2023 → Kinderanteile enden 2023, übrige Kennzahlen bis 2025."),
 ("Datenqualität: '_2019'-Datensätze","Trotz Namenssuffix '_2019' enthalten staatsangehoerigkeit/haushalte_personen Zeitreihen bis 2025."),
 ("Kinderbetreuungsplätze","Vorhanden auf Stadtbezirksebene, jedoch im Datensatz 'sozialmonitoring-betreuungsplaetze-fur-kinder' (nicht stadtbezirke_*): Plätze <3 J. und 3–6 J., 2015–2025. Als Versorgungsquote NICHT vorberechnet."),
 ("Aktuellstes Jahr je Kennzahl","Spaltenüberschriften in 'Konsolidiert' tragen das jeweils verwendete Jahr in [eckigen Klammern]; 'Zeitreihe' enthält die volle Historie."),
 ("Nicht aufgenommen (separat dokumentiert)","Detail-Breakdowns: auslaender_altersgruppen, *_staatsangehoerigkeit_2019, haushalte_personen_2019, *_altersgruppen_weiblich, hauptwohnsitz_altersgruppen_weiblich – vorhanden im Portal, hier nicht flachgejoint."),
]
ws5.append(["Würzburg Stadtbezirke – Quellen & Methodik"]); ws5.cell(1,1).font=Font(bold=True,size=12); ws5.append([])
for k,v in notes:
    ws5.append([k,v])
    rr=ws5.max_row
    ws5.cell(rr,1).font=Font(bold=True); ws5.cell(rr,1).alignment=Alignment(vertical="top")
    ws5.cell(rr,2).alignment=Alignment(vertical="top",wrap_text=True)
ws5.column_dimensions["A"].width=30; ws5.column_dimensions["B"].width=95

OUT="Wuerzburg_Stadtbezirke_Sozialdaten.xlsx"
wb.save(OUT)
print("WROTE",OUT)
print("Sheets:",wb.sheetnames)
PY_END=1
